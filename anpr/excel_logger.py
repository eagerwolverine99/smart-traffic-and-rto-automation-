"""
Excel Logger for ANPR System
=============================
Writes detected plates to a 3-sheet Excel workbook:
  1. Detections  — every confirmed plate reading with full info
  2. Unique      — deduplicated (one row per unique plate)
  3. Summary     — counts, state-wise breakdown

Features:
  - Auto-saves after each write (crash-safe)
  - Embeds plate snapshot thumbnail in the Excel cell
  - Saves individual plate snapshots to disk as well
  - Uses openpyxl (pure Python, no Excel needed)

Installation:
    pip install openpyxl pillow
"""

import os
import cv2
import time
from datetime import datetime
from collections import Counter
from io import BytesIO

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARNING] openpyxl not installed. Run: pip install openpyxl pillow")

from plate_info import decode_plate


# =====================================================================
# STYLE CONSTANTS
# =====================================================================
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78',
                          fill_type='solid') if OPENPYXL_AVAILABLE else None
HEADER_FONT = Font(color='FFFFFF', bold=True,
                   size=11) if OPENPYXL_AVAILABLE else None
HEADER_ALIGN = Alignment(horizontal='center', vertical='center',
                         wrap_text=True) if OPENPYXL_AVAILABLE else None
CELL_ALIGN = Alignment(horizontal='left', vertical='center',
                       wrap_text=False) if OPENPYXL_AVAILABLE else None

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
) if OPENPYXL_AVAILABLE else None


# =====================================================================
# DETECTIONS SHEET COLUMNS
# =====================================================================
DETECTION_COLUMNS = [
    ('Sr. No.',       6),
    ('Timestamp',     20),
    ('Track ID',      10),
    ('Plate Number',  18),
    ('State',         22),
    ('RTO Code',      12),
    ('City / Region', 28),
    ('Category',      28),
    ('Series',        15),
    ('Vehicle Type',  14),
    ('Confidence',    12),
    ('OCR Reads',     12),
    ('Plate Image',   22),
]

UNIQUE_COLUMNS = [
    ('Sr. No.',       6),
    ('Plate Number',  18),
    ('State',         22),
    ('RTO / City',    30),
    ('Category',      28),
    ('Vehicle Type',  14),
    ('First Seen',    20),
    ('Last Seen',     20),
    ('Times Seen',    12),
    ('Best Conf.',    12),
]


# =====================================================================
# EXCEL LOGGER CLASS
# =====================================================================
class ExcelLogger:
    def __init__(self, output_dir='output', filename='anpr_log.xlsx'):
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl pillow")

        self.output_dir = output_dir
        self.snapshots_dir = os.path.join(output_dir, 'plate_snapshots')
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(self.snapshots_dir, exist_ok=True)

        self.filepath = os.path.join(output_dir, filename)

        # State
        self.detections = []   # list of detection dicts
        self.unique_plates = {}  # plate_text -> dict

        # Build workbook
        self.wb = Workbook()
        self._setup_sheets()
        self._save()

    def _setup_sheets(self):
        # Remove default sheet and create our three
        self.wb.remove(self.wb.active)

        self.ws_det = self.wb.create_sheet('Detections', 0)
        self.ws_uniq = self.wb.create_sheet('Unique Plates', 1)
        self.ws_sum = self.wb.create_sheet('Summary', 2)

        self._setup_detections_sheet()
        self._setup_unique_sheet()
        self._setup_summary_sheet()

    def _setup_detections_sheet(self):
        ws = self.ws_det
        for col_idx, (name, width) in enumerate(DETECTION_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30
        # Freeze header row
        ws.freeze_panes = 'A2'

    def _setup_unique_sheet(self):
        ws = self.ws_uniq
        for col_idx, (name, width) in enumerate(UNIQUE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

    def _setup_summary_sheet(self):
        ws = self.ws_sum
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.cell(row=1, column=1, value='ANPR Session Summary')
        ws.cell(row=1, column=1).font = Font(size=16, bold=True,
                                             color='1F4E78')

    # ----------------------------------------------------------------
    def log_plate(self, plate_text, track_id, confidence, num_readings,
                  plate_img=None, vehicle_class_id=None):
        """
        Log a confirmed plate detection.

        Args:
            plate_text: e.g. "KA02MN1826"
            track_id: tracker ID for this vehicle
            confidence: 0-1 float
            num_readings: number of OCR reads that voted
            plate_img: np array of the plate crop (BGR), optional
            vehicle_class_id: YOLO class ID (2=car, 3=bike, etc.)
        """
        timestamp = datetime.now()
        timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')

        # Decode plate info
        info = decode_plate(plate_text, vehicle_class_id)

        # Save snapshot to disk
        snapshot_path = None
        if plate_img is not None and plate_img.size > 0:
            safe_name = plate_text.replace('/', '_')
            filename = f"{safe_name}_{timestamp.strftime('%Y%m%d_%H%M%S')}.jpg"
            snapshot_path = os.path.join(self.snapshots_dir, filename)
            try:
                cv2.imwrite(snapshot_path, plate_img)
            except Exception:
                snapshot_path = None

        # Build detection record
        record = {
            'timestamp': timestamp_str,
            'track_id': track_id,
            'plate': plate_text,
            'state': info['state'],
            'rto_code': info['rto_code'],
            'city': info['city'],
            'category': info['category'],
            'series': info['series'],
            'vehicle_type': info['vehicle_type'],
            'confidence': confidence,
            'num_readings': num_readings,
            'snapshot_path': snapshot_path,
        }

        self.detections.append(record)

        # Update unique plates
        if plate_text in self.unique_plates:
            u = self.unique_plates[plate_text]
            u['last_seen'] = timestamp_str
            u['times_seen'] += 1
            if confidence > u['best_conf']:
                u['best_conf'] = confidence
        else:
            self.unique_plates[plate_text] = {
                'plate': plate_text,
                'state': info['state'],
                'rto_city': f"{info['rto_code']} ({info['city']})",
                'category': info['category'],
                'vehicle_type': info['vehicle_type'],
                'first_seen': timestamp_str,
                'last_seen': timestamp_str,
                'times_seen': 1,
                'best_conf': confidence,
            }

        # Write to sheets
        self._append_detection_row(record)
        self._rebuild_unique_sheet()
        self._rebuild_summary_sheet()
        self._save()

    # ----------------------------------------------------------------
    def _append_detection_row(self, rec):
        ws = self.ws_det
        row = ws.max_row + 1
        sr_no = len(self.detections)

        values = [
            sr_no,
            rec['timestamp'],
            rec['track_id'],
            rec['plate'],
            rec['state'],
            rec['rto_code'],
            rec['city'],
            rec['category'],
            rec['series'],
            rec['vehicle_type'],
            f"{rec['confidence']:.1%}",
            rec['num_readings'],
            '',  # image placeholder
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

        # Embed snapshot image
        if rec.get('snapshot_path') and os.path.exists(rec['snapshot_path']):
            try:
                img = XLImage(rec['snapshot_path'])
                # Resize for spreadsheet cell
                img.width = 140
                img.height = 42
                col_letter = get_column_letter(len(DETECTION_COLUMNS))
                img.anchor = f'{col_letter}{row}'
                ws.add_image(img)
                ws.row_dimensions[row].height = 36
            except Exception:
                pass

    # ----------------------------------------------------------------
    def _rebuild_unique_sheet(self):
        ws = self.ws_uniq
        # Clear existing data rows (keep header)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Sort by first_seen
        sorted_plates = sorted(
            self.unique_plates.values(),
            key=lambda p: p['first_seen']
        )

        for i, p in enumerate(sorted_plates, 1):
            row = i + 1
            values = [
                i,
                p['plate'],
                p['state'],
                p['rto_city'],
                p['category'],
                p['vehicle_type'],
                p['first_seen'],
                p['last_seen'],
                p['times_seen'],
                f"{p['best_conf']:.1%}",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.alignment = CELL_ALIGN
                cell.border = THIN_BORDER

    # ----------------------------------------------------------------
    def _rebuild_summary_sheet(self):
        ws = self.ws_sum
        # Clear existing (except title)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        row = 3

        def add_row(label, value, bold=False):
            nonlocal row
            cell_a = ws.cell(row=row, column=1, value=label)
            cell_b = ws.cell(row=row, column=2, value=value)
            if bold:
                cell_a.font = Font(bold=True)
                cell_b.font = Font(bold=True)
            row += 1

        add_row('Generated at:',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 1

        add_row('OVERVIEW', '', bold=True)
        add_row('Total detections:', len(self.detections))
        add_row('Unique plates:', len(self.unique_plates))
        row += 1

        # State-wise breakdown
        state_counts = Counter(p['state'] for p in self.unique_plates.values())
        if state_counts:
            add_row('STATE-WISE BREAKDOWN', '', bold=True)
            for state, count in state_counts.most_common():
                add_row(f'  {state}', count)
            row += 1

        # Vehicle type breakdown
        type_counts = Counter(p['vehicle_type']
                              for p in self.unique_plates.values())
        if type_counts:
            add_row('VEHICLE TYPE BREAKDOWN', '', bold=True)
            for vtype, count in type_counts.most_common():
                add_row(f'  {vtype}', count)
            row += 1

        # Category breakdown
        cat_counts = Counter(p['category']
                             for p in self.unique_plates.values())
        if cat_counts:
            add_row('PLATE CATEGORY', '', bold=True)
            for cat, count in cat_counts.most_common():
                add_row(f'  {cat}', count)

    # ----------------------------------------------------------------
    def _save(self):
        try:
            self.wb.save(self.filepath)
        except PermissionError:
            print(f"[WARNING] Could not save {self.filepath} — "
                  f"is it open in Excel?")
        except Exception as e:
            print(f"[WARNING] Save failed: {e}")

    def close(self):
        self._save()
        print(f"\n[EXCEL] Log saved to: {self.filepath}")
        print(f"[EXCEL] Snapshots: {self.snapshots_dir}")
        print(f"[EXCEL] {len(self.detections)} detections, "
              f"{len(self.unique_plates)} unique plates")
